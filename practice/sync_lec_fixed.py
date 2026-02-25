import openpyxl
import re

excel_path = r'C:\DataMajor\practice\leetcode_files.xlsx'
md_path = r'C:\DataMajor\practice\001Study\LEC.md'

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read LEC.md table first
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    table_lines = re.findall(r'^\|\s*\*\*(\d+)\*\*\s*\|\s*\[([^\]]+)\]\([^)]+\)\s*\|\s*(.*?)\s*\|\s*(\d+)/10\s*\|', content, re.MULTILINE)
    
    # Create dictionary of problems from LEC.md
    lec_problems = {}
    for match in table_lines:
        num = str(match[0].strip())
        title = match[1].strip()
        concepts_raw = match[2]
        difficulty = int(match[3].strip())
        
        concepts = [c.strip(' `') for c in concepts_raw.split(',')]
        # pad to 5
        while len(concepts) < 5: concepts.append(None)
        
        lec_problems[num] = {
            'title': title,
            'difficulty': difficulty,
            'concepts': concepts
        }

    # Clean up mistakes and gather existing proper rows
    rows_to_delete = []
    existing_nums = set()
    
    for i in range(ws.max_row, 1, -1):
        worked_on = ws.cell(row=i, column=1).value
        prob_num = ws.cell(row=i, column=3).value
        # If I messed up and wrote '{num}.ipynb' into the WorkedOn column earlier, let's delete that row
        if type(worked_on) == str and worked_on.endswith('.ipynb'):
            ws.delete_rows(i)
        elif prob_num is not None:
            existing_nums.add(str(prob_num).strip())

    added_count = 0
    updated_count = 0

    # Add missing problems from LEC.md
    for num, data in lec_problems.items():
        if num not in existing_nums:
            new_row = [
                'x',                                     # 0: WorkedOn
                f"LC{int(num):04d}",                     # 1: LeetCode Number
                str(num),                                # 2: Problem Number
                data['title'],                           # 3: File Name (Putting title for now)
                r"C:\DataMajor\practice\001Study\LEC.md",# 4: Full Path
                data['difficulty'],                      # 5: Difficulty
                data['concepts'][0],                     # 6: Concept 1
                data['concepts'][1],                     # 7: Concept 2
                data['concepts'][2],                     # 8: Concept 3
                data['concepts'][3],                     # 9: Concept 4
                data['concepts'][4]                      # 10: Concept 5
            ]
            ws.append(new_row)
            added_count += 1
        else:
            # Update existing problems with the correct difficulty and concepts
            for row in ws.iter_rows(min_row=2):
                if str(row[2].value).strip() == str(num):
                    row[0].value = 'x' # mark as worked on since it's in LEC.md
                    # Check if diff needs update
                    row[5].value = data['difficulty']
                    row[6].value = data['concepts'][0]
                    row[7].value = data['concepts'][1]
                    row[8].value = data['concepts'][2]
                    row[9].value = data['concepts'][3]
                    row[10].value = data['concepts'][4]
                    updated_count += 1
                    break

    wb.save(excel_path)
    print(f"Cleanup and Sync complete! Added {added_count} missing rows, updated {updated_count} existing rows with latest LEC.md data.")

except Exception as e:
    print(f"Error: {e}")
