"""
build_leetcode_excel.py
Reads leetcode_files.csv, fills Difficulty + Concepts, writes leetcode_files.xlsx
Difficulty scale: 1 (easiest) → 10 (hardest), relative to the full LC spectrum.
"""

import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Problem metadata  key = integer problem number
# (difficulty, [concept1, concept2, concept3, concept4, concept5])
# ---------------------------------------------------------------------------
PROBLEM_META = {
    1:   (2,  ["Hash Map", "Array", "One Pass", "", ""]),
    3:   (5,  ["Sliding Window", "Hash Map", "String", "Two Pointers", ""]),
    11:  (5,  ["Two Pointers", "Array", "Greedy", "", ""]),
    15:  (6,  ["Two Pointers", "Array", "Sorting", "Hash Map", ""]),
    20:  (3,  ["Stack", "String", "Hash Map", "", ""]),
    21:  (3,  ["Linked List", "Two Pointers", "Recursion", "", ""]),
    33:  (6,  ["Binary Search", "Array", "Divide & Conquer", "", ""]),
    48:  (6,  ["Array", "Matrix", "In-Place", "Math", ""]),
    49:  (5,  ["Hash Map", "String", "Sorting", "", ""]),
    53:  (4,  ["Dynamic Programming", "Kadane's Algorithm", "Array", "Divide & Conquer", ""]),
    54:  (6,  ["Array", "Matrix", "Simulation", "", ""]),
    56:  (6,  ["Array", "Sorting", "Intervals", "", ""]),
    68:  (9,  ["String", "Simulation", "Greedy", "", ""]),
    70:  (2,  ["Dynamic Programming", "Math", "Memoization", "", ""]),
    73:  (5,  ["Array", "Matrix", "Hash Map", "", ""]),
    74:  (5,  ["Binary Search", "Matrix", "Array", "", ""]),
    76:  (8,  ["Sliding Window", "Hash Map", "String", "Two Pointers", ""]),
    84:  (8,  ["Stack", "Monotonic Stack", "Array", "", ""]),
    100: (2,  ["Tree", "DFS", "Recursion", "", ""]),
    102: (4,  ["Tree", "BFS", "Queue", "", ""]),
    104: (2,  ["Tree", "DFS", "Recursion", "BFS", ""]),
    121: (2,  ["Array", "Greedy", "One Pass", "", ""]),
    128: (5,  ["Hash Set", "Array", "Union Find", "", ""]),
    141: (3,  ["Linked List", "Two Pointers", "Floyd's Algorithm", "", ""]),
    146: (7,  ["Hash Map", "Doubly Linked List", "Design", "OrderedDict", ""]),
    150: (5,  ["Stack", "Array", "Math", "", ""]),
    153: (5,  ["Binary Search", "Array", "", "", ""]),
    154: (7,  ["Binary Search", "Array", "Duplicates", "", ""]),
    155: (4,  ["Stack", "Design", "Auxiliary Stack", "", ""]),
    167: (3,  ["Two Pointers", "Array", "Binary Search", "", ""]),
    198: (4,  ["Dynamic Programming", "Array", "", "", ""]),
    200: (5,  ["DFS", "BFS", "Matrix", "Union Find", ""]),
    206: (2,  ["Linked List", "Iterative", "Recursion", "", ""]),
    209: (5,  ["Sliding Window", "Array", "Two Pointers", "", ""]),
    217: (1,  ["Hash Set", "Array", "", "", ""]),
    226: (2,  ["Tree", "DFS", "Recursion", "BFS", ""]),
    227: (7,  ["Stack", "String", "Math", "", ""]),
    238: (6,  ["Array", "Prefix Sum", "Suffix Product", "", ""]),
    242: (2,  ["Hash Map", "String", "Frequency Count", "Sorting", ""]),
    347: (5,  ["Hash Map", "Heap", "Bucket Sort", "", ""]),
    378: (2,  ["Hash Map", "String", "Frequency Count", "", ""]),   # as used in this project (first-unique-char)
    387: (2,  ["Hash Map", "String", "Queue", "", ""]),
    424: (6,  ["Sliding Window", "Hash Map", "String", "Two Pointers", ""]),
    438: (5,  ["Sliding Window", "Hash Map", "String", "", ""]),
    496: (4,  ["Stack", "Monotonic Stack", "Hash Map", "", ""]),
    503: (5,  ["Stack", "Monotonic Stack", "Circular Array", "", ""]),
    523: (6,  ["Hash Map", "Prefix Sum", "Math", "", ""]),
    525: (5,  ["Hash Map", "Prefix Sum", "Array", "", ""]),
    560: (5,  ["Hash Map", "Prefix Sum", "Array", "", ""]),
    567: (5,  ["Sliding Window", "Hash Map", "String", "", ""]),
    674: (4,  ["Dynamic Programming", "Array", "Greedy", "", ""]),
    704: (1,  ["Binary Search", "Array", "", "", ""]),
    739: (5,  ["Stack", "Monotonic Stack", "Array", "", ""]),
    901: (6,  ["Stack", "Monotonic Stack", "Design", "", ""]),
    937: (4,  ["String", "Sorting", "Custom Sort", "", ""]),
    974: (6,  ["Hash Map", "Prefix Sum", "Math", "", ""]),
    1004:(5,  ["Sliding Window", "Array", "Two Pointers", "", ""]),
    1861:(6,  ["Array", "Matrix", "Two Pointers", "Simulation", ""]),
    2043:(4,  ["Array", "Design", "Simulation", "", ""]),
    3045:(8,  ["String", "Hashing", "Z-Algorithm", "Trie", ""]),
    3161:(9,  ["Binary Indexed Tree", "Segment Tree", "Binary Search", "", ""]),
}

# Difficulty colour bands  (fill hex)
def diff_fill(d):
    if d <= 2:   return "C6EFCE"   # green
    if d <= 4:   return "FFEB9C"   # yellow
    if d <= 6:   return "FFCC99"   # orange
    if d <= 8:   return "FF9999"   # red-light
    return           "CC0000"      # dark red  (9-10)

def diff_font_color(d):
    return "9C0006" if d >= 9 else "000000"

# ---------------------------------------------------------------------------
# Read CSV
# ---------------------------------------------------------------------------
csv_path  = r"C:\DataMajor\practice\leetcode_files.csv"
xlsx_path = r"C:\DataMajor\practice\leetcode_files.xlsx"

rows = []
with open(csv_path, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    fieldnames = reader.fieldnames
    for row in reader:
        rows.append(dict(row))

# ---------------------------------------------------------------------------
# Enrich each row
# ---------------------------------------------------------------------------
for row in rows:
    raw = row.get("Problem Number", "").strip()
    try:
        num = int(raw)
    except ValueError:
        num = -1

    if num in PROBLEM_META:
        diff, concepts = PROBLEM_META[num]
        row["Difficulty"] = diff
        row["Concept 1"]  = concepts[0]
        row["Concept 2"]  = concepts[1]
        row["Concept 3"]  = concepts[2]
        row["Concept 4"]  = concepts[3]
        row["Concept 5"]  = concepts[4]
    # else leave blanks as-is

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "LeetCode Files"

HEADERS = ["LeetCode Number", "Problem Number", "File Name",
           "Full Path", "Difficulty",
           "Concept 1", "Concept 2", "Concept 3", "Concept 4", "Concept 5"]

# --- Header row styling ---
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
center      = Alignment(horizontal="center", vertical="center", wrap_text=False)
thin        = Side(style="thin", color="AAAAAA")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADERS)
for col_idx, _ in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = center
    cell.border = border

# --- Data rows ---
for r_idx, row in enumerate(rows, 2):
    vals = [row.get(h, "") for h in HEADERS]
    ws.append(vals)

    diff_val = row.get("Difficulty", "")
    for c_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.border    = border
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if c_idx in (1,2,5) else "left")

        # Colour Difficulty cell
        if c_idx == 5 and isinstance(diff_val, int):
            cell.fill = PatternFill("solid", fgColor=diff_fill(diff_val))
            cell.font = Font(bold=True, color=diff_font_color(diff_val))

# --- Column widths ---
col_widths = [14, 14, 52, 80, 12, 22, 22, 22, 22, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row height
for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
    ws.row_dimensions[row_cells[0].row].height = 16

# Freeze panes & autofilter
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ---------------------------------------------------------------------------
wb.save(xlsx_path)
print(f"Saved: {xlsx_path}  ({len(rows)} data rows)")
