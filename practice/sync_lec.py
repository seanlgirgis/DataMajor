import openpyxl
import re
import os

excel_path = r'C:\DataMajor\practice\leetcode_files.xlsx'
md_path = r'C:\DataMajor\practice\001Study\LEC.md'

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get existing problem numbers
    existing_nums = set()
    for row in ws.iter_rows(min_row=2):
        if row[2].value is not None:
            existing_nums.add(str(row[2].value).strip())

    # Read LEC.md
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the table part
    # Look for lines starting with | **<num>** |
    table_lines = re.findall(r'^\|\s*\*\*(\d+)\*\*\s*\|\s*\[([^\]]+)\]\([^)]+\)\s*\|\s*(.*?)\s*\|\s*(\d+)/10\s*\|', content, re.MULTILINE)
    
    added_count = 0
    updated_count = 0
    
    for match in table_lines:
        num = match[0].strip()
        title = match[1].strip()
        concepts_raw = match[2]
        difficulty = int(match[3].strip())
        
        # Parse concepts (they are in backticks, e.g., `Concept 1`, `Concept 2`)
        concepts = [c.strip(' `') for c in concepts_raw.split(',')]
        
        c1 = concepts[0] if len(concepts) > 0 else None
        c2 = concepts[1] if len(concepts) > 1 else None
        c3 = concepts[2] if len(concepts) > 2 else None
        c4 = concepts[3] if len(concepts) > 3 else None
        c5 = concepts[4] if len(concepts) > 4 else None

        if num not in existing_nums:
            # Add new row
            # ['File Name', 'Folder Path', 'Prob num', 'Problem Name', 'Difficulty (1-10)', 'Difficulty Level', 'Concept 1', 'Concept 2', 'Concept 3', 'Concept 4', 'Concept 5']
            new_row = [
                f"{num}.ipynb",  # Fake a file name or leave blank? Let's use standard naming or blank. Let's use 'LEC.md' as file name
                r"C:\DataMajor\practice\001Study",
                int(num),
                title,
                None,  # Difficulty (1-10) is column 4. Wait, usually difficulty level is 1,2,3...
                difficulty,
                c1,
                c2,
                c3,
                c4,
                c5
            ]
            ws.append(new_row)
            added_count += 1
            print(f"Added Problem {num}: {title}")
        else:
            # Update existing row
            for row in ws.iter_rows(min_row=2):
                if str(row[2].value).strip() == str(num):
                    # update concepts and difficulty
                    row[5].value = difficulty
                    if c1: row[6].value = c1
                    if c2: row[7].value = c2
                    if c3: row[8].value = c3
                    if c4: row[9].value = c4
                    if c5: row[10].value = c5
                    updated_count += 1
                    break

    wb.save(excel_path)
    print(f"Successfully synced. Added {added_count} new problems, updated {updated_count} existing problems.")

except Exception as e:
    print(f"Error: {e}")
